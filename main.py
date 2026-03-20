import requests, os, re, csv, io, xlrd, gc, json
import pandas as pd, geopandas as gpd, numpy as np
from charset_normalizer import from_path
from concurrent.futures import ThreadPoolExecutor
import unicodedata, inflection
from pymongo import MongoClient
from openpyxl import load_workbook
from dotenv import load_dotenv
import logging

logging.basicConfig(
    filename="recife.log",
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    force=True
)

load_dotenv()

PRIORITY = {
    "csv": 0,
    "json": 1,
    "xlsx": 2,
    "geojson": 3,
}

STOPWORDS_PT = [
    "O","A","OS","AS","UM","UMA","UNS","UMAS",
    "ANTE","APOS","ATE","COM","CONTRA","DE","DESDE","EM","ENTRE",
    "PARA","PERANTE","POR","SEM","SOB","SOBRE","TRAS",
    "DO","DA","DOS","DAS","DUM","DUMA","DUNS","DUMAS",
    "DELE","DELA","DELES","DELAS","DESTE","DESTA","DESTES","DESTAS",
    "DESSE","DESSA","DESSES","DESSAS",
    "DAQUELE","DAQUELA","DAQUELES","DAQUELAS",
    "NO","NA","NOS","NAS","NUM","NUMA","NUNS","NUMAS",
    "NELE","NELA","NELES","NELAS",
    "NESTE","NESTA","NESTES","NESTAS",
    "NESSE","NESSA","NESSES","NESSAS",
    "NAQUELE","NAQUELA","NAQUELES","NAQUELAS",
    "AO","AOS",
    "PELO","PELA","PELOS","PELAS",
    "E","OU","MAS","QUE","SE","COMO", 
    "º"
]

STOPWORDS_REPLACE = {
    "TRIMESTRE": "TRIM",
    "TRIMESTRES": "TRIM",
}

STOPWORDS = {s: '' for s in STOPWORDS_PT}
STOPWORDS.update(STOPWORDS_REPLACE)
STOPWORDS_KEYS = list(STOPWORDS.keys())

DOWNLOAD_FOLDER = './temp'

HOST = os.getenv('HOST')
DATABASE_NAME = 'DATASETS_RECIFE'


def get_datasets_list() -> list[str]:

    try:

        logging.info('Obtendo lista de datasets disponíveis...')

        response = requests.get('https://dados.recife.pe.gov.br/api/3/action/package_list')
        response.raise_for_status()

        response_list = response.json().get('result', [])
        
        logging.info(f'Total de datasets encontrados: {len(response_list)}')
        
        return response_list

    except Exception as e:
        logging.error(f'Erro ao listar os datasets | {str(e)}')

    return []


def get_dataset_details(dataset_name) -> dict:
    try:


        response = requests.get(f'https://dados.recife.pe.gov.br/api/3/action/package_search?q={dataset_name}')
        response.raise_for_status()

        metadatas = response.json().get('result', {}).get('results', [])
        
        details = []

        for metadata in metadatas:
            resources = metadata.get('resources', [])

            tags = [tag.get('name', '') for tag in metadata.get('tags', [])]

            for resource in resources:
                details.append({
                    'dataset_name': dataset_name,
                    'grupo_name': metadata.get('name', ''),
                    'notes': metadata.get('notes', ''),
                    'organizacao_title': metadata.get('organization', {}).get('title', ''),
                    'organizacao_description': metadata.get('organization', {}).get('description', ''),
                    'resource_name': resource.get('name', ''),
                    'url': resource.get('url', ''),
                    'resource_description': resource.get('description', ''),
                    'format': resource.get('format', ''),
                    'last_modified': resource.get('last_modified', ''),
                    'tags': tags
                })        

        #details = delete_duplicate_datasets(details)

        return details
        
    except Exception as e:
        logging.error(f'Erro ao obter detalhes do dataset | {str(e)}')

    return None


def delete_duplicate_datasets(details: list[dict]) -> list[dict]:
    try:
        df = pd.DataFrame(details)
        df["last_modified"] = pd.to_datetime(df["last_modified"])
            
        df = df.sort_values("last_modified", ascending=False).drop_duplicates("url")

        df["priority"] = df["format"].str.lower()
        df["format"] = df["format"].str.lower()
        df = df[df["format"].isin(PRIORITY)]
        df["priority"] = df["format"].map(PRIORITY)

        df = df.sort_values("priority").drop_duplicates("url").drop(columns="priority")

        details = df.to_dict(orient='records')

    except Exception as e:
        logging.error(f'Erro ao eliminar datasets duplicados | {str(e)}')

    return details


def remove_acentos(texto: str) -> str:
    """Remove acentos de uma string usando unicodedata."""
    nfkd = unicodedata.normalize('NFD', texto)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def delete_dictionary(details: list[dict]) -> list[dict]:
    try:

        df = pd.DataFrame(details)
        mask = df['resource_name'].apply(lambda x: 'dicionario' not in remove_acentos(str(x)).lower())
        df = df[mask]
        return df.to_dict(orient='records')

    except Exception as e:
        logging.error(f'Erro ao eliminar datasets duplicados | {str(e)}')
    
    return None


def delete_metadatas(details: list[dict]) -> list[dict]:
    """Elimina arquivos de metadados."""
    try:
        df = pd.DataFrame(details)
        # Filtra metadados que possuem arquivos e formatos válidos
        #mask = (df['url'].notnull()) & (df['format'].notnull()) & (df['format'] != '')
        mask = df['resource_name'].apply(lambda x: 'metadados' not in remove_acentos(str(x)).lower())
        
        df = df[mask]
        return df.to_dict(orient='records')
    except Exception as e:
        logging.error(f'Erro ao eliminar metadados inválidos | {str(e)}')
    return None


def delete_anexos(details: list[dict]) -> list[dict]:
    """Elimina arquivos de anexos."""
    try:
        df = pd.DataFrame(details)
        mask = df['resource_name'].apply(lambda x: 'anexos' not in remove_acentos(str(x)).lower())
        df = df[mask]
        return df.to_dict(orient='records')
    except Exception as e:
        logging.error(f'Erro ao eliminar anexos inválidos | {str(e)}')
    return None


def get_collection_name(resource_name:str) -> str:
    name = resource_name.upper()
    name = remove_acentos(name)
    name = re.sub(r'[^a-zA-Z0-9 _-]', '', name)
    name = name.replace(' ', '_')
    name = name.replace('-', '_')
    
    t_s = name.split('_')
    t_s = [STOPWORDS[w] if w in STOPWORDS_KEYS else w for w in t_s]
    name = "_".join(t for t in t_s if t)

    name = re.sub(r"_+", "_", name)

    return name


def download_file(url, resource_name:str, format:str) -> str | None:
    try:

        response = requests.get(url, stream=True)
        collection_name = get_collection_name(resource_name)
        file_path = os.path.join(DOWNLOAD_FOLDER, collection_name) + f'.{format}'

        with open(file_path, "wb") as f:
            f.write(response.content)

        if os.path.exists(file_path):
            return file_path

    except Exception as e:
        logging.error(f'Erro ao baixar o arquivo | {str(e)}')

    return None


def get_encoding_csv(file_path: str) -> str | None:
    """
    Detecta encoding do CSV usando charset-normalizer
    com fallback manual.
    """
    try:
        result = from_path(file_path).best()
        if result and result.encoding:
            return result.encoding
    except Exception:
        pass

    encodings_to_try = ["utf-8", "utf-8-sig", "latin1", "cp1252"]

    for encoding in encodings_to_try:
        try:
            with open(file_path, encoding=encoding) as f:
                f.readline()
            return encoding
        except UnicodeDecodeError:
            continue
        except Exception:
            return None

    return None


def get_delimiter_csv(file_path: str, encoding: str) -> str | None:
    """
    Detecta delimiter usando Sniffer com fallback heurístico.
    """

    try:
        with open(file_path, encoding=encoding) as f:
            sample = f.read(10000)

            if not sample.strip():
                return None

            dialect = csv.Sniffer().sniff(sample)
            return dialect.delimiter

    except Exception:
        pass

    delimiters = [",", ";", "|", "\t"]
    counts = {d: [] for d in delimiters}

    try:
        with open(file_path, encoding=encoding) as f:
            for _ in range(20):
                line = f.readline()
                if not line:
                    break

                for d in delimiters:
                    counts[d].append(line.count(d))

        scores = {}

        for d, values in counts.items():
            if not values:
                continue

            avg = sum(values) / len(values)
            variance = max(values) - min(values)

            if avg > 0:
                scores[d] = (avg, variance)

        if not scores:
            return None

        # prioriza maior média e menor variância
        best = sorted(scores.items(), key=lambda x: (-x[1][0], x[1][1]))[0][0]
        return best

    except Exception:
        return None


def detect_header_offset(csv_path, encoding, delimiter):

    with open(csv_path, encoding=encoding) as f:
        for i, line in enumerate(f):
            if line.count(delimiter) >= 2:  # pelo menos 3 colunas
                if next((False for t in line.split(delimiter) if not t), True): # detecta se o nome da coluna não for vazia
                    return i

    return 0


def column_normalizer(s: str):
    s = s.replace('\n', '').strip()
    s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore')
    s = s.decode("ASCII")
    s = s.strip().replace(' ', '_').replace('.', '_').replace('\"', '')
    s = inflection.underscore(s).upper()
    s = re.sub('[^A-Za-z0-9]+', '_', s)

    s = re.sub(r"_+", "_", s)
      
    if s.endswith("JSON"): s = s.replace("JSON", "")
    if s.endswith("CSV"): s = s.replace("CSV", "")
    if s.endswith("_"): s = s[:-1]
    if s.startswith("_"): s = s[1:]
    return s


def make_unique_columns(columns):
    seen = {}
    new_cols = []

    for col in columns:
        if col not in seen:
            seen[col] = 0
            new_cols.append(col)
        else:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")

    return new_cols


def sanitized_file_like(file_path: str, encoding: str):
    with open(file_path, "r", encoding=encoding, errors="ignore") as f:
        content = f.read().replace("\x00", "")
    return io.StringIO(content)


def infer_column_type(series: pd.Series) -> pd.Series:
    s = series.dropna()

    if s.empty:
        return series

    # Tenta inteiro
    try:
        converted = pd.to_numeric(s, errors="raise", downcast="integer")
        if not converted.isna().any():
            return pd.to_numeric(series, errors="coerce").astype("Int64")
    except:
        pass

    # Tenta float
    try:
        converted = pd.to_numeric(s, errors="raise")
        if not converted.isna().any():
            return pd.to_numeric(series, errors="coerce")
    except:
        pass

    # Tenta boolean
    lower = s.str.lower()
    if lower.isin(["true", "false"]).all():
        return series.str.lower().map({"true": True, "false": False})

    # Caso contrário mantém string
    return series


def insert_into_database(collection_name:str, data: list[dict]) -> bool:
    try:

        if not data:
            return False
        
        df = pd.DataFrame(data)
        df.replace(np.nan, None, inplace=True)
        df.fillna('None', inplace=True)
        df.replace('None', None, inplace=True)
        data = df.to_dict(orient='records')

        if not collection_name.endswith('_exec'):
            collection_name = collection_name + '_exec'

        with MongoClient(HOST) as client:
            collection = client[DATABASE_NAME][collection_name]
            collection.insert_many(data)

        return True

    except Exception as e:
        logging.error(f'Erro ao inserir no banco de dados | {str(e)}')

    return False


def drop_and_rename_collection_exec(collection_name:str) -> bool:
    try:

        if collection_name.endswith('_exec'):
            collection_name = collection_name[:-5]

        with MongoClient(HOST) as client:
            db = client[DATABASE_NAME]
            if collection_name + '_exec' in db.list_collection_names():
                if collection_name in db.list_collection_names():
                    db[collection_name].drop()

                db[collection_name + '_exec'].rename(collection_name)
                return True

    except Exception as e:
        logging.error(f'Erro ao renomear coleção no banco de dados | {str(e)}')

    return False


def excel_to_csv_safe(file_path: str, output_path: str | None = None) -> str | None:
    """
    Converte XLS/XLSX para CSV de forma segura (Windows-safe).
    """

    wb = None
    workbook = None

    try:
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in (".xlsx", ".xls"):
            return None

        if output_path is None:
            base_path = os.path.splitext(file_path)[0]
            output_path = f"{base_path}.csv"

        output_dir = os.path.dirname(output_path) or "."
        os.makedirs(output_dir, exist_ok=True)

        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:

            writer = csv.writer(
                f,
                delimiter=";",
                quotechar='"',
                quoting=csv.QUOTE_MINIMAL
            )

            if ext == ".xlsx":

                wb = load_workbook(file_path, read_only=True)
                ws = wb.active

                for row in ws.iter_rows(values_only=True):
                    cleaned_row = [
                        str(cell).replace("\n", " ")
                        .replace("\r", " ")
                        .replace('"', "'")
                        if cell is not None else ""
                        for cell in row
                    ]
                    writer.writerow(cleaned_row)

                wb.close()
                wb = None


            elif ext == ".xls":

                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_index(0)

                for i in range(sheet.nrows):
                    row = sheet.row_values(i)

                    cleaned_row = [
                        str(cell).replace("\n", " ")
                        .replace("\r", " ")
                        .replace('"', "'")
                        if cell is not None else ""
                        for cell in row
                    ]

                    writer.writerow(cleaned_row)

                del workbook
                workbook = None


        gc.collect()


        # agora pode remover
        if os.path.abspath(file_path) != os.path.abspath(output_path):
            os.remove(file_path)

        return output_path


    except Exception as e:

        logging.error(f"Erro ao converter Excel para CSV | {e}")

        if output_path and os.path.exists(output_path):
            os.remove(output_path)

        return None


def extrair_lista_registros(obj):
    if isinstance(obj, list) and all(isinstance(i, dict) for i in obj):
        return obj

    # Se for dict, procurar listas dentro dele
    if isinstance(obj, dict):
        for value in obj.values():
            result = extrair_lista_registros(value)
            if result:
                return result
    
    return None


def processing_and_insert_file(file_path:str, chunk_size:int = 50_000, add_columns:dict = None) -> list[dict]:
    """Processa o arquivo, detectando tipo, limpando e inserindo no banco de dados."""

    inserted = False
    collection_name = os.path.splitext(os.path.basename(file_path))[0]

    try:
        
        if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
            converted_path = excel_to_csv_safe(file_path)
            if not converted_path:
                return False
            
            file_path = converted_path


        if file_path.endswith('.csv') or file_path.endswith('.geojson'):
            encoding = get_encoding_csv(file_path)

            if not encoding:
                return False


        if file_path.endswith('.csv'):

            delimiter = get_delimiter_csv(file_path, encoding)

            if not delimiter:
                return False

            offset = detect_header_offset(file_path, encoding, delimiter)

            header = pd.read_csv(
                sanitized_file_like(file_path, encoding),
                delimiter=delimiter,
                encoding=encoding,
                nrows=0,
                skiprows=offset,
            ).columns

            normalized_columns = [column_normalizer(col) for col in header]
            normalized_columns = make_unique_columns(normalized_columns)

            # lê em chunks para arquivos grandes, aplicando limpeza e inferência de tipos
            chunks = pd.read_csv(
                sanitized_file_like(file_path, encoding),
                delimiter=delimiter,
                encoding=encoding,
                engine="python",
                on_bad_lines='skip',
                chunksize=chunk_size,
                #quoting=csv.QUOTE_NONE, # ignora aspas // isso quebra o pipelines
                skiprows=offset + 1,
                dtype=str,
                header=None,
                names=normalized_columns,
            )

            try:
                chunk = next(chunks)
                chunk = chunk.apply(lambda col: col.str.strip('"') if col.dtype == "object" else col)

                if chunk.empty:
                    return False
            except Exception as e:
                return False
            
            for column in chunk.columns:
                chunk[column] = infer_column_type(chunk[column])

            schema = chunk.dtypes.to_dict()

            while chunk is not None and not chunk.empty:

                for column, dtype in schema.items():

                    if 'unnamed' in column.lower():
                        chunk.drop(column, axis=1, inplace=True)
                    else:
                        try:
                            chunk[column] = chunk[column].astype(dtype)
                        except:
                            pass  # mantém string se falhar
                    
                chunk = chunk.dropna(how="all")

                if not chunk.empty:
                    if add_columns: # adiciona colunas extras
                        chunk = chunk.assign(**dict(add_columns))

                    if insert_into_database(collection_name, chunk.to_dict('records')):
                        inserted = True

                try:
                    chunk = next(chunks)
                    chunk = chunk.apply(lambda col: col.str.strip('"') if col.dtype == "object" else col)
                except:
                    break


        elif file_path.endswith('.json'):
            
            with open(file_path, 'r') as f:
                json_data = json.load(f)

            data = extrair_lista_registros(json_data)
            if not data:
                return False
            
            for i in range(0, len(data), chunk_size):
                chunk = pd.DataFrame(data[i:i+chunk_size])

                if chunk.empty:
                    continue

                chunk.columns = [column_normalizer(col) for col in chunk.columns]

                if add_columns:
                    chunk = chunk.assign(**dict(add_columns))

                if insert_into_database(collection_name, chunk.to_dict('records')):
                    inserted = True


        elif file_path.endswith('.geojson'):
            gdf = gpd.read_file(
                file_path,
                encoding=encoding            
            )
            gdf.columns = [column_normalizer(col) for col in gdf.columns]
            gdf['GEOMETRY'] = gdf['GEOMETRY'].apply(lambda g: g.__geo_interface__).tolist()
            
            data = gdf.to_dict('records')

            for i in range(0, len(data), chunk_size): 
                chunk = pd.DataFrame(data[i:i+chunk_size])

                if chunk.empty:
                    continue
                
                if insert_into_database(collection_name, gdf.to_dict('records')):
                    inserted = True

    except Exception as e:
        logging.error(f'Erro ao processar o arquivo | {str(e)}')

    finally:
        if inserted:
            drop_and_rename_collection_exec(collection_name)
            logging.info(f'Arquivo processado e inserido com sucesso: {collection_name}')

        if os.path.exists(file_path):
            os.remove(file_path)

    return None


def run():
    try:

        logging.info('----------> Iniciando rotina de ingestão de datasets do Recife...')

        dataset_list = get_datasets_list()
        
        details = []
        futures = []

        with ThreadPoolExecutor(max_workers=5) as executor:
            for dataset_name in dataset_list: 
                future = executor.submit(get_dataset_details, dataset_name)
                futures.append(future)


            for future in futures:
                dataset_details = future.result()

                if dataset_details:
                    details.extend(dataset_details)

        logging.info(f'Total de recursos encontrados: {len(details)}')

        
        #df = pd.DataFrame(details)
        #df['format'].unique()
        #df[df['format'].str.lower() == 'geojson']
        #df[df['format'].str.lower() == 'geojson'].iloc[1]['url']

        logging.info('Eliminando recursos duplicados, dicionários, metadados e anexos...')

        details = delete_duplicate_datasets(details)
        details = delete_dictionary(details)
        details = delete_metadatas(details)
        details = delete_anexos(details)

        logging.info(f'Total de recursos após limpeza: {len(details)}')


        def download_process_insert(detail):
            url = detail.get('url', '')
            resource_name = detail.get('resource_name', '')
            format = detail.get('format', '').lower()

            file_path = download_file(url, resource_name, format)
            if file_path:
                processing_and_insert_file(file_path)

        logging.info('----------> Iniciando download, processamento e inserção dos arquivos...')
        
        
        with ThreadPoolExecutor(max_workers=5) as executor:
            for detail in details:
                executor.submit(download_process_insert, detail)
        
        #for detail in details: 
        #    url = detail.get('url', '')
        #    resource_name = detail.get('resource_name', '')
        #    format = detail.get('format', '').lower()
        #
        #    file_path = download_file(url, resource_name, format)
        #    if file_path:
        #        processing_and_insert_file(file_path)

    except Exception as e:
        logging.error(f'Erro na rotina principal | {str(e)}')

if __name__ == '__main__':
    run()